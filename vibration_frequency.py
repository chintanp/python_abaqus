# Save by Administrator on 2015_07_29-11.30.48; build 6.12-1 2012_03_13-20.44.39 119612
from part import *
from material import *
from section import *
from assembly import *
from step import *
from interaction import *
from load import *
from mesh import *
from optimization import *
from job import *
from sketch import *
from visualization import *
from connectorBehavior import *

import os 
# change the abaqus working directory
os.chdir('E://rahul')

# import openpyxl for excel manupulation

from openpyxl import load_workbook
wb = load_workbook(filename = "E://rahul//inputdynamic.xlsx")
ws = wb.active

for j in range(0, len(ws.rows)): 
    #pass
    width_beam  = float(ws.rows[j][0].value)#float(ws['B].value) # 200.
    height_beam = float(ws.rows[j][1].value) # 20.
    length_beam = float(ws.rows[j][2].value) # 25.
    nel = float(ws.rows[j][3].value)



    mdb.models['Model-1'].ConstrainedSketch(name='__profile__', sheetSize=200.0)

    mdb.models['Model-1'].sketches['__profile__'].rectangle(point1=(0.0, 0.0), 
        point2= (width_beam, height_beam))
    mdb.models['Model-1'].Part(dimensionality=THREE_D, name='Part-1', type=
        DEFORMABLE_BODY)
    mdb.models['Model-1'].parts['Part-1'].BaseSolidExtrude(depth=length_beam, sketch=
        mdb.models['Model-1'].sketches['__profile__'])
    del mdb.models['Model-1'].sketches['__profile__']
    mdb.models['Model-1'].Material(name='Material-1')
    mdb.models['Model-1'].materials['Material-1'].Density(table=((1000.0, ), ))
    mdb.models['Model-1'].materials['Material-1'].Elastic(table=((100000000000.0, 
        0.2), ))
    mdb.models['Model-1'].HomogeneousSolidSection(material='Material-1', name=
        'Section-1', thickness=None)
    mdb.models['Model-1'].parts['Part-1'].SectionAssignment(offset=0.0, 
        offsetField='', offsetType=MIDDLE_SURFACE, region=Region(
        cells=mdb.models['Model-1'].parts['Part-1'].cells.getSequenceFromMask(
        mask=('[#1 ]', ), )), sectionName='Section-1', thicknessAssignment=
        FROM_SECTION)
    mdb.models['Model-1'].rootAssembly.DatumCsysByDefault(CARTESIAN)
    mdb.models['Model-1'].rootAssembly.Instance(dependent=ON, name='Part-1-1', 
        part=mdb.models['Model-1'].parts['Part-1'])
    mdb.models['Model-1'].FrequencyStep(name='Step-1', numEigen=12, previous=
        'Initial')
    mdb.models['Model-1'].parts['Part-1'].seedPart(deviationFactor=0.1, 
        minSizeFactor=0.1, size=(length_beam/nel))
    mdb.models['Model-1'].parts['Part-1'].generateMesh()
    mdb.models['Model-1'].parts['Part-1'].setElementType(elemTypes=(ElemType(
        elemCode=C3D8R, elemLibrary=STANDARD, secondOrderAccuracy=OFF, 
        kinematicSplit=AVERAGE_STRAIN, hourglassControl=DEFAULT, 
        distortionControl=DEFAULT), ElemType(elemCode=C3D6, elemLibrary=STANDARD), 
        ElemType(elemCode=C3D4, elemLibrary=STANDARD)), regions=(
        mdb.models['Model-1'].parts['Part-1'].cells.getSequenceFromMask(('[#1 ]', 
        ), ), ))
    mdb.models['Model-1'].rootAssembly.regenerate()
    mdb.models['Model-1'].EncastreBC(createStepName='Step-1', localCsys=None, name=
        'BC-1', region=Region(
        faces=mdb.models['Model-1'].rootAssembly.instances['Part-1-1'].faces.getSequenceFromMask(
        mask=('[#10 ]', ), )))
    mdb.Job(atTime=None, contactPrint=OFF, description='', echoPrint=OFF, 
        explicitPrecision=SINGLE, getMemoryFromAnalysis=True, historyPrint=OFF, 
        memory=50, memoryUnits=PERCENTAGE, model='Model-1', modelPrint=OFF, 
        multiprocessingMode=DEFAULT, name='Job_dyn', nodalOutputPrecision=SINGLE, 
        numCpus=1, numGPUs=0, queue=None, scratch='', type=ANALYSIS, 
        userSubroutine='', waitHours=0, waitMinutes=0)
    mdb.jobs['Job_dyn'].submit(consistencyChecking=OFF)
    mdb.jobs['Job_dyn'].waitForCompletion()


    from openpyxl import Workbook
    wbo = Workbook()
    wso = wbo.create_sheet(0)



    from odbAccess import *
    odb = openOdb(path='vibration.odb')
    step1 = odb.steps['Step-1']
    region = step1.historyRegions
    region = step1.historyRegions['Assembly ASSEMBLY']
    freqData = region.historyOutputs['EIGFREQ'].data

    for i in range(0, len(freqData)):
        wso['A' + str(i+1)].value = float(freqData[i][0])
        wso['B' + str(i+1)].value = float(freqData[i][1])


    wbo.save("E://Vibration//output_dynamic"+ str(j) + ".xlsx")


